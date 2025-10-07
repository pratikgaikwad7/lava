from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session, make_response
from admin_app import get_db_connection
from datetime import datetime, timedelta, date, time
from utils import Config, Constants, load_training_data
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from collections import defaultdict
import pandas as pd
import os
from flask import current_app
import pymysql.cursors
import pymysql

# Blueprint definition
view_bp = Blueprint('view_bp', __name__)

# Constants
RECORDS_PER_PAGE = 200

# Helper function to check if user is logged in
def is_logged_in():
    return 'logged_in' in session and session['logged_in']

# Helper function to check if user has specific role
def has_role(role_name):
    return is_logged_in() and session.get('role') == role_name

# Helper function to get current user info
def get_current_user():
    if is_logged_in():
        return {
            'id': session.get('user_id'),
            'username': session.get('username'),
            'role': session.get('role'),
            'factory_location': session.get('factory_location')
        }
    return None

# Helper function to apply factory filter based on user's role and factory location
def apply_user_factory_filter(filters):
    """Apply factory filter based on user's role and factory location"""
    # Get user role and factory location from session
    user_role = session.get('role', '')
    user_factory = session.get('factory_location', '')
    
    # Factory-specific roles
    factory_specific_roles = ["Factory Head", "PSD", "Shop Floor Training Coordinators"]
    
    # If user has a factory-specific role, override factory filter
    if user_role in factory_specific_roles and user_factory:
        filters['factory'] = user_factory
    
    return filters

# Login route
@view_bp.route('/login', methods=['GET', 'POST'])
def login():
    # If user is already logged in, redirect to master data
    if is_logged_in():
        return redirect(url_for('view_bp.view_master_data'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        factory_location = request.form.get('factory_location')
        
        # Validate required fields
        if not username or not password or not role:
            flash("Please fill in all required fields", "error")
            return render_template('user_tech/login.html')
            
        # Check if factory location is required but not provided
        if role in ["Factory Head", "PSD", "Shop Floor Training Coordinators"] and not factory_location:
            flash("Factory Location is required for this role", "error")
            return render_template('user_tech/login.html')
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed", "error")
            return render_template('user_tech/login.html')
            
        try:
            with conn.cursor(pymysql.cursors.DictCursor) as cursor:
                # Check if user exists with the provided credentials
                if role in ["Factory Head", "PSD", "Shop Floor Training Coordinators"]:
                    # For factory-specific roles, we need to check factory location too
                    cursor.execute(
                        "SELECT * FROM user_auth WHERE username = %s AND password = %s AND role = %s AND factory_location = %s",
                        (username, password, role, factory_location)
                    )
                else:
                    # For other roles
                    cursor.execute(
                        "SELECT * FROM user_auth WHERE username = %s AND password = %s AND role = %s",
                        (username, password, role)
                    )
                    
                user = cursor.fetchone()
                
                if user:
                    # Set session variables
                    session['user_id'] = user['id']
                    session['username'] = user['username']
                    session['role'] = user['role']
                    session['factory_location'] = user['factory_location']
                    session['logged_in'] = True
                    session['login_time'] = datetime.now().isoformat()
                    session.permanent = True
                    
                    # Redirect to master data page
                    return redirect(url_for('view_bp.view_master_data'))
                else:
                    flash("Invalid credentials. Please try again.", "error")
                    return render_template('user_tech/login.html')
                    
        except Exception as e:
            flash(f"An error occurred: {str(e)}", "error")
            return render_template('user_tech/login.html')
        finally:
            if conn:
                conn.close()
    
    # For GET requests, show login form with cache control headers
    response = make_response(render_template('user_tech/login.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# Logout route
@view_bp.route('/logout')
def logout():
    # Clear all session data
    session.clear()
    
    # Create response with cache control headers
    response = make_response(redirect(url_for('view_bp.login')))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    
    flash("You have been logged out successfully", "success")
    return response

# Before request handler to check authentication and set cache control
@view_bp.before_request
def require_login():
    # List of endpoints that don't require authentication
    allowed_endpoints = ['view_bp.login', 'view_bp.logout', 'static']
    
    # Check if the request endpoint is in the allowed routes
    if request.endpoint and request.endpoint in allowed_endpoints:
        # Set cache control headers for login page
        if request.endpoint == 'view_bp.login':
            # We'll set headers in the response after the view function
            pass
        return
    
    # Redirect to login if not logged in
    if not is_logged_in():
        # Create response with cache control headers
        response = make_response(redirect(url_for('view_bp.login')))
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
        return response

# After request handler to set cache control headers for all responses
@view_bp.after_request
def add_cache_headers(response):
    # Skip cache control for static files
    if request.endpoint and request.endpoint == 'static':
        return response
    
    # Set cache control headers to prevent caching
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    
    return response

# Helper Functions
def calculate_learning_hours(record):
    """Calculate learning hours - use manual value if exists, otherwise calculate based on attendance"""
    manual_hours = record.get('learning_hours')
    if manual_hours is not None and manual_hours != '' and manual_hours != 0:
        return int(manual_hours)
    
    program_hours = int(record.get('program_hours', 0))
    day1 = bool(record.get('day_1_attendance'))
    day2 = bool(record.get('day_2_attendance'))
    day3 = bool(record.get('day_3_attendance'))
    
    if program_hours <= 8:
        return program_hours if day1 else 0
    else:
        attended_days = sum([1 for day in [day1, day2, day3] if day])
        return min(attended_days * 8, program_hours)

def get_fiscal_year(date=None, return_string=False):
    """Get fiscal year (April-March) for a given date or current date.
    If return_string is True, returns formatted string (e.g., "FY 2025-26").
    Otherwise, returns the fiscal year as an integer (e.g., 2025).
    """
    if date is None:
        date = datetime.now()
    year = date.year if date.month >= 4 else date.year - 1
    
    if return_string:
        next_year_short = str(year + 1)[-2:]
        return f"FY {year}-{next_year_short}"
    else:
        return year

def get_fiscal_year_range(year):
    """Get start and end dates for a fiscal year (April 1 - March 31)"""
    # If year is a string in format "FY 2025-26", extract the integer part
    if isinstance(year, str) and year.startswith("FY "):
        year = int(year.split()[1].split('-')[0])
    
    start_date = datetime(year, 4, 1).date()
    end_date = datetime(year + 1, 3, 31).date()
    return start_date, end_date

def parse_date(date_val):
    """Parse date string or date object to datetime.date object"""
    if not date_val:
        return None
    
    # If it's already a date object, return it
    if isinstance(date_val, date):
        return date_val
    
    # If it's a string, try to parse it
    formats = [
        '%Y-%m-%d',    # 2025-07-16
        '%d-%m-%Y',    # 03-04-2025
        '%d/%m/%Y',    # 03/04/2025
        '%m/%d/%Y',    # 04/03/2025 (US format)
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_val, fmt).date()
        except (ValueError, TypeError):
            continue
    
    return None

def format_date(date_val):
    """Format date string or date object to DD/MM/YYYY"""
    date_obj = parse_date(date_val)
    if date_obj:
        return date_obj.strftime('%d/%m/%Y')
    return date_val

def format_time(time_val):
    """Format time string or time object to HH:MM"""
    if not time_val:
        return ""
    
    # If it's already a time object, format it
    if isinstance(time_val, time):
        return time_val.strftime('%H:%M')
    
    # If it's a string, try to parse it
    try:
        time_obj = datetime.strptime(time_val, '%H:%M:%S').time()
        return time_obj.strftime('%H:%M')
    except (ValueError, TypeError):
        try:
            time_obj = datetime.strptime(time_val, '%H:%M').time()
            return time_obj.strftime('%H:%M')
        except (ValueError, TypeError):
            return time_val

# Filter Application Functions
def apply_fiscal_year_filter(query, params, fiscal_year):
    """Apply fiscal year filter to a query"""
    if not fiscal_year:
        return query, params
    
    # If fiscal_year is a string in format "FY 2025-26", extract the integer part
    if isinstance(fiscal_year, str) and fiscal_year.startswith("FY "):
        fiscal_year = int(fiscal_year.split()[1].split('-')[0])
    else:
        fiscal_year = int(fiscal_year)
    
    query += """
        AND (
            (start_date IS NOT NULL AND 
             (YEAR(start_date) + IF(MONTH(start_date) >= 4, 0, -1)) = %s)
        )
    """
    params.extend([fiscal_year])
    return query, params

def apply_date_range_filter(query, params, start_date_str, end_date_str):
    """Apply date range filter to a query"""
    if not start_date_str or not end_date_str:
        return query, params
    
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    if not start_date or not end_date:
        return query, params
    
    query += """
        AND (
            (start_date >= %s AND end_date <= %s)
        )
    """
    params.extend([start_date, end_date])
    return query, params

def apply_month_range_filter(query, params, month_range_start, month_range_end):
    """Apply month range filter to a query"""
    if not month_range_start or not month_range_end:
        return query, params
    
    month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                 'July', 'August', 'September', 'October', 'November', 'December']
    try:
        start_idx = month_order.index(month_range_start)
        end_idx = month_order.index(month_range_end)
        
        if start_idx > end_idx:
            months_in_range = month_order[start_idx:] + month_order[:end_idx+1]
        else:
            months_in_range = month_order[start_idx:end_idx+1]
            
        placeholders = ','.join(['%s'] * len(months_in_range))
        query += f" AND calendar_month IN ({placeholders})"
        params.extend(months_in_range)
    except ValueError:
        pass
    
    return query, params

def apply_pmo_training_category_filter(query, params, pmo_training_category):
    """Apply PMO training category filter to a query"""
    if not pmo_training_category or pmo_training_category == 'All':
        return query, params
    
    if pmo_training_category == 'PMO':
        query += " AND pmo_training_category != 'SHE (Safety+Health)'"
    else:
        query += " AND pmo_training_category = %s"
        params.append(pmo_training_category)
    
    return query, params

def apply_pl_category_filter(query, params, pl_category):
    """Apply PL category filter to a query"""
    if not pl_category or pl_category == 'All':
        return query, params
    
    query += " AND pl_category = %s"
    params.append(pl_category)
    return query, params

def apply_standard_filters(query, params, filters):
    """Apply standard filters to a query"""
    # Apply fiscal year filter
    query, params = apply_fiscal_year_filter(query, params, filters.get('fiscal_year'))
    
    # Apply other standard filters
    if filters.get('per_no'):
        query += " AND per_no = %s"
        params.append(filters['per_no'])
    
    if filters.get('bc_no'):
        query += " AND bc_no = %s"
        params.append(filters['bc_no'])
    
    if filters.get('gender') and filters['gender'] != 'All':
        query += " AND gender = %s"
        params.append(filters['gender'])
    
    if filters.get('calendar_month'):
        query += " AND calendar_month = %s"
        params.append(filters['calendar_month'])
    
    if filters.get('month_report_pmo_21_20'):
        query += " AND month_report_pmo_21_20 = %s"
        params.append(filters['month_report_pmo_21_20'])
    
    if filters.get('month_cd_key_26_25'):
        query += " AND month_cd_key_26_25 = %s"
        params.append(filters['month_cd_key_26_25'])
    
    # FIXED: Changed TNI status filter condition
    if filters.get('tni_status'):
        query += " AND tni_non_tni = %s"
        params.append(filters['tni_status'])
    
    if filters.get('training_name'):
        query += " AND training_name = %s"
        params.append(filters['training_name'])
    
    if filters.get('employee_group'):
        query += " AND employee_group = %s"
        params.append(filters['employee_group'])
    
    if filters.get('factory'):
        query += " AND factory = %s"
        params.append(filters['factory'])
    
    # Apply date range filter
    query, params = apply_date_range_filter(query, params, 
                                          filters.get('start_date'), 
                                          filters.get('end_date'))
    
    # Apply PL category filter
    query, params = apply_pl_category_filter(query, params, filters.get('pl_category'))
    
    # Apply PMO training category filter
    query, params = apply_pmo_training_category_filter(query, params, 
                                                     filters.get('pmo_training_category'))
    
    # Apply month range filter
    query, params = apply_month_range_filter(query, params, 
                                           filters.get('month_range_start'), 
                                           filters.get('month_range_end'))
    
    # Apply user factory filter based on role
    filters = apply_user_factory_filter(filters)
    
    return query, params

def build_base_query(filters, for_export=False):
    """Build the base SQL query with filters"""
    # For export, we don't need the id/sr_no column
    if for_export:
        base_query = """
            SELECT 
                per_no,
                participants_name,
                bc_no,
                gender,
                employee_group,
                department,
                factory,
                training_name,
                pmo_training_category,
                pl_category,
                brsr_sq_123_category,
                calendar_need_base_reschedule as program_type,
                tni_non_tni as tni_status,
                learning_hours,
                start_date,
                end_date,
                calendar_month,
                month_report_pmo_21_20,
                month_cd_key_26_25,
                start_time,
                end_time,
                location_hall,
                faculty_1,
                faculty_2,
                faculty_3,
                faculty_4,
                mobile_no,
                email,
                Nomination_received_from,
                cordi_name as verified_by,
                day_1_attendance,
                day_2_attendance,
                day_3_attendance,
                (day_1_attendance + IFNULL(day_2_attendance, 0) + IFNULL(day_3_attendance, 0)) as attended_days,
                learning_hours as program_hours
            FROM master_data
            WHERE 1=1
        """
    else:
        base_query = """
            SELECT 
                id as sr_no,
                per_no,
                participants_name,
                bc_no,
                gender,
                employee_group,
                department,
                factory,
                training_name,
                pmo_training_category,
                pl_category,
                brsr_sq_123_category,
                calendar_need_base_reschedule as program_type,
                tni_non_tni as tni_status,
                learning_hours,
                start_date,
                end_date,
                calendar_month,
                month_report_pmo_21_20,
                month_cd_key_26_25,
                start_time,
                end_time,
                location_hall,
                faculty_1,
                faculty_2,
                faculty_3,
                faculty_4,
                mobile_no,
                email,
                Nomination_received_from,
                cordi_name as verified_by,
                day_1_attendance,
                day_2_attendance,
                day_3_attendance,
                (day_1_attendance + IFNULL(day_2_attendance, 0) + IFNULL(day_3_attendance, 0)) as attended_days,
                learning_hours as program_hours
            FROM master_data
            WHERE 1=1
        """
    
    query_params = []
    
    # Apply all standard filters
    base_query, query_params = apply_standard_filters(base_query, query_params, filters)
    
    base_query += " ORDER BY id DESC"
    
    # Only apply pagination for web view, not for exports
    if not for_export:
        page = request.args.get('page', 1, type=int)
        offset = (page - 1) * RECORDS_PER_PAGE
        base_query += f" LIMIT {RECORDS_PER_PAGE} OFFSET {offset}"
    
    return base_query, query_params

def get_column_headings():
    """Return the column headings for both HTML and Excel in the specified order"""
    return {
        'sr_no': 'SR.No',
        'per_no': 'Per. No',
        'participants_name': 'Participants Name',
        'bc_no': 'BC No',
        'gender': 'Gender',
        'employee_group': 'Employee group',
        'department': 'Department',
        'factory': 'Factory',
        'training_name': 'Training Name',
        'pmo_training_category': 'PMO Training Category',
        'pl_category': 'PL Category',
        'brsr_sq_123_category': 'BRSR SQ 1,2,3 Category',
        'program_type': 'Calendar/Need base/Reschedule',
        'tni_status': 'TNI / NON TNI',
        'learning_hours': 'Learning hours',
        'start_date': 'Start Date (DD/MM/YYYY)',
        'end_date': 'End Date (DD/MM/YYYY)',
        'calendar_month': 'Calendar Month',
        'month_report_pmo_21_20': 'Month Report PMO 21-20',
        'month_cd_key_26_25': 'Month CD/Key 26-25',
        'start_time': 'Start Time',
        'end_time': 'End Time',
        'location_hall': 'Location Hall',
        'faculty_1': 'Faculty 1',
        'faculty_2': 'Faculty 2',
        'faculty_3': 'Faculty 3',
        'faculty_4': 'Faculty 4',
        'mobile_no': 'Mobile No',
        'email': 'Email',
        'Nomination_received_from': 'Nomination Received From',
        'verified_by': 'Cordinator Name',
        'day_1_attendance': 'Day 1',
        'day_2_attendance': 'Day 2',
        'day_3_attendance': 'Day 3',
        'attended_days': 'Days Attended'
    }

def process_records(raw_records):
    """Process raw database records for display"""
    records = []
    for record in raw_records:
        processed_record = dict(record)
        
        day1 = bool(processed_record.get('day_1_attendance'))
        day2 = bool(processed_record.get('day_2_attendance'))
        day3 = bool(processed_record.get('day_3_attendance'))
        
        processed_record['learning_hours'] = calculate_learning_hours(processed_record)
        processed_record['start_date'] = format_date(processed_record.get('start_date'))
        processed_record['end_date'] = format_date(processed_record.get('end_date'))
        processed_record['start_time'] = format_time(processed_record.get('start_time'))
        processed_record['end_time'] = format_time(processed_record.get('end_time'))
        processed_record['day_1_attendance'] = 'Yes' if day1 else 'No'
        processed_record['day_2_attendance'] = 'Yes' if day2 else 'No'
        processed_record['day_3_attendance'] = 'Yes' if day3 else 'No'
        
        records.append(processed_record)
    return records

def clean_training_name(training_name):
    """Clean training name by removing metadata in quotes"""
    if not training_name:
        return ""
    
    # Remove everything after the first single quote if it exists
    parts = training_name.split("'", 1)
    return parts[0].strip()

@view_bp.route('/get_training_names')
def get_training_names():
    """Endpoint to fetch training names from training_names table (all available training programs)"""
    conn = get_db_connection()
    if not conn:
        return jsonify([])
    
    try:
        with conn.cursor() as cursor:
            # Get all training programs from the training_names table
            cursor.execute("SELECT DISTINCT Training_Name FROM training_names ORDER BY Training_Name")
            training_names = [row['Training_Name'] for row in cursor.fetchall()]
            return jsonify(training_names)
    except Exception as e:
        print(f"Error fetching training names: {str(e)}")
        return jsonify([])
    finally:
        if conn:
            conn.close()

@view_bp.route('/get_training_programs')
def get_training_programs():
    """Endpoint to fetch metadata about all available training programs"""
    conn = get_db_connection()
    if not conn:
        return jsonify([])
    
    try:
        with conn.cursor() as cursor:
            # Get all training programs without filtering by TNI status
            cursor.execute("""
                SELECT Training_Name, PMO_Training_Category, PL_Category, 
                       BRSR_SQ_123_Category, Tni_Status, learning_hours
                FROM training_names
                ORDER BY Training_Name
            """)
            training_data = cursor.fetchall()
            return jsonify(training_data)
    except Exception as e:
        print(f"Error fetching training programs: {str(e)}")
        return jsonify([])
    finally:
        if conn:
            conn.close()

def calculate_dashboard_metrics(filters):
    """Calculate dashboard metrics based on filters"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        # First get the total count of all records (based on id)
        count_query = "SELECT COUNT(id) as total_records FROM master_data WHERE 1=1"
        query_params = []
        
        # Apply all filters to the count query
        count_query, query_params = apply_standard_filters(count_query, query_params, filters)
        
        # Execute the count query
        with conn.cursor() as cursor:
            cursor.execute(count_query, query_params)
            count_result = cursor.fetchone()
            total_records = count_result['total_records'] if count_result else 0
            
        # Now get the actual records for other calculations using the same filters
        base_query, base_query_params = build_base_query(filters, for_export=True)
        
        with conn.cursor() as cursor:
            cursor.execute(base_query, base_query_params)
            all_records = cursor.fetchall()
            
            learning_hours = 0
            for record in all_records:
                record_dict = dict(record)
                learning_hours += calculate_learning_hours(record_dict)
                
            # Get count of unique permanent learners with PER_NO
            unique_query = """
                SELECT COUNT(DISTINCT per_no) as unique_permanent_learners 
                FROM master_data 
                WHERE employee_group = 'PERMANENT' 
                AND per_no IS NOT NULL
                AND per_no != ''
            """
            unique_params = []
            
            # Apply ALL filters from the base query to ensure consistency
            unique_query, unique_params = apply_standard_filters(unique_query, unique_params, filters)
            
            # Execute unique learner query
            cursor.execute(unique_query, unique_params)
            unique_result = cursor.fetchone()
            unique_learners = unique_result['unique_permanent_learners'] if unique_result else 0
            
            # Calculate target metrics with consistent PMO/SHE filtering and fiscal year
            target_query = """
                SELECT 
                    SUM(t.hours) as target_hours,
                    COUNT(DISTINCT t.per_no) as target_unique_learners,
                    COUNT(*) as target
                FROM final_tni_data t
                JOIN training_targets tt ON t.training_name = tt.training_name
                WHERE 1=1
            """
            target_params = []
            
            # Apply fiscal year filter to target metrics
            if filters.get('fiscal_year'):
                fiscal_year = int(filters['fiscal_year'])
                target_query += " AND t.year = %s AND tt.target_year = %s"
                target_params.extend([fiscal_year, fiscal_year])
            
            if filters.get('factory'):
                target_query += " AND t.factory = %s"
                target_params.append(filters['factory'])
            
            if filters.get('training_name'):
                target_query += " AND t.training_name = %s"
                target_params.append(filters['training_name'])
            
            if filters.get('bc_no'):
                target_query += " AND t.bc_no = %s"
                target_params.append(filters['bc_no'])
            
            # Apply PMO/SHE filter to target metrics
            if filters.get('pmo_training_category'):
                if filters['pmo_training_category'] == 'PMO':
                    target_query += " AND tt.pmo_category != 'SHE (Safety+Health)'"
                elif filters['pmo_training_category'] != 'All':
                    target_query += " AND tt.pmo_category = %s"
                    target_params.append(filters['pmo_training_category'])
            
            if filters.get('pl_category') and filters['pl_category'] != 'All':
                target_query += " AND tt.pl_category = %s"
                target_params.append(filters['pl_category'])
            
            cursor.execute(target_query, target_params)
            target_result = cursor.fetchone()
            
            target_metrics = {
                'target_hours': target_result['target_hours'] if target_result and target_result['target_hours'] else 0,
                'target_unique_learners': target_result['target_unique_learners'] if target_result and target_result['target_unique_learners'] else 0,
                'target': target_result['target'] if target_result and target_result['target'] else 0
            }
            
            # Calculate TNI data metrics with consistent PMO/SHE filtering and fiscal year
            tni_query = """
                SELECT 
                    COUNT(DISTINCT per_no) as tni_unique_learners,
                    COUNT(*) as tni_total_count
                FROM tni_data
                WHERE 1=1
            """
            tni_params = []
            
            # Apply fiscal year filter to TNI metrics
            if filters.get('fiscal_year'):
                fiscal_year = int(filters['fiscal_year'])
                tni_query += " AND year = %s"
                tni_params.append(fiscal_year)
            
            if filters.get('factory'):
                tni_query += " AND factory = %s"
                tni_params.append(filters['factory'])
            
            if filters.get('training_name'):
                tni_query += " AND training_name = %s"
                tni_params.append(filters['training_name'])
            
            if filters.get('bc_no'):
                tni_query += " AND bc_no = %s"
                tni_params.append(filters['bc_no'])
            
            # Apply PMO/SHE filter to TNI metrics
            if filters.get('pmo_training_category'):
                if filters['pmo_training_category'] == 'PMO':
                    tni_query += " AND training_name IN (SELECT training_name FROM training_targets WHERE pmo_category != 'SHE (Safety+Health)'"
                    if filters.get('fiscal_year'):
                        tni_query += " AND target_year = %s"
                        tni_params.append(fiscal_year)
                    tni_query += ")"
                elif filters['pmo_training_category'] != 'All':
                    tni_query += " AND training_name IN (SELECT training_name FROM training_targets WHERE pmo_category = %s"
                    tni_params.append(filters['pmo_training_category'])
                    if filters.get('fiscal_year'):
                        tni_query += " AND target_year = %s"
                        tni_params.append(fiscal_year)
                    tni_query += ")"
            
            if filters.get('pl_category') and filters['pl_category'] != 'All':
                tni_query += " AND training_name IN (SELECT training_name FROM training_targets WHERE pl_category = %s"
                tni_params.append(filters['pl_category'])
                if filters.get('fiscal_year'):
                    tni_query += " AND target_year = %s"
                    tni_params.append(fiscal_year)
                tni_query += ")"
            
            cursor.execute(tni_query, tni_params)
            tni_result = cursor.fetchone()
            tni_metrics = {
                'tni_total_count': tni_result['tni_total_count'] if tni_result else 0,
                'tni_unique_learners': tni_result['tni_unique_learners'] if tni_result else 0
            }
            
            # Calculate matching records between TNI data and actual attendance with consistent filtering
            if tni_metrics['tni_total_count'] > 0:
                match_query = """
                    SELECT COUNT(DISTINCT t.id) as matched_count
                    FROM tni_data t
                    JOIN master_data m ON t.per_no = m.per_no
                        AND t.factory = m.factory
                        AND t.training_name = m.training_name
                    WHERE 1=1
                """
                match_params = []
                
                # Apply fiscal year filter to match query
                if filters.get('fiscal_year'):
                    fiscal_year = int(filters['fiscal_year'])
                    
                    # For master_data, we need to derive fiscal year from start_date
                    match_query += """
                        AND (
                            (m.start_date IS NOT NULL AND 
                             (YEAR(m.start_date) + IF(MONTH(m.start_date) >= 4, 0, -1)) = %s)
                        )
                        AND t.year = %s
                    """
                    match_params.extend([fiscal_year, fiscal_year])
                
                if filters.get('factory'):
                    match_query += " AND t.factory = %s"
                    match_params.append(filters['factory'])
                
                if filters.get('training_name'):
                    match_query += " AND t.training_name = %s"
                    match_params.append(filters['training_name'])
                
                if filters.get('bc_no'):
                    match_query += " AND t.bc_no = %s"
                    match_params.append(filters['bc_no'])
                
                if filters.get('employee_group'):
                    match_query += " AND m.employee_group = %s"
                    match_params.append(filters['employee_group'])
                
                # Apply PMO/SHE filter to match query
                if filters.get('pmo_training_category'):
                    if filters['pmo_training_category'] == 'PMO':
                        match_query += " AND t.training_name IN (SELECT training_name FROM training_targets WHERE pmo_category != 'SHE (Safety+Health)'"
                        if filters.get('fiscal_year'):
                            match_query += " AND target_year = %s"
                            match_params.append(fiscal_year)
                        match_query += ")"
                    elif filters['pmo_training_category'] != 'All':
                        match_query += " AND t.training_name IN (SELECT training_name FROM training_targets WHERE pmo_category = %s"
                        match_params.append(filters['pmo_training_category'])
                        if filters.get('fiscal_year'):
                            match_query += " AND target_year = %s"
                            match_params.append(fiscal_year)
                        match_query += ")"
                
                if filters.get('pl_category') and filters['pl_category'] != 'All':
                    match_query += " AND t.training_name IN (SELECT training_name FROM training_targets WHERE pl_category = %s"
                    match_params.append(filters['pl_category'])
                    if filters.get('fiscal_year'):
                        match_query += " AND target_year = %s"
                        match_params.append(fiscal_year)
                    match_query += ")"
                
                cursor.execute(match_query, match_params)
                match_result = cursor.fetchone()
                matched_count = match_result['matched_count'] if match_result else 0
                
                # Count of distinct TNI records without matching master_data records
                remaining_query = """
                    SELECT COUNT(DISTINCT t.id) as remaining_count
                    FROM tni_data t
                    LEFT JOIN master_data m ON t.per_no = m.per_no
                        AND t.factory = m.factory
                        AND t.training_name = m.training_name
                    WHERE m.per_no IS NULL
                """
                remaining_params = []
                
                # Apply fiscal year filter to remaining query
                if filters.get('fiscal_year'):
                    fiscal_year = int(filters['fiscal_year'])
                    remaining_query += " AND t.year = %s"
                    remaining_params.append(fiscal_year)
                
                if filters.get('factory'):
                    remaining_query += " AND t.factory = %s"
                    remaining_params.append(filters['factory'])
                
                if filters.get('training_name'):
                    remaining_query += " AND t.training_name = %s"
                    remaining_params.append(filters['training_name'])
                
                if filters.get('bc_no'):
                    remaining_query += " AND t.bc_no = %s"
                    remaining_params.append(filters['bc_no'])
                
                # Apply PMO/SHE filter to remaining query
                if filters.get('pmo_training_category'):
                    if filters['pmo_training_category'] == 'PMO':
                        remaining_query += " AND t.training_name IN (SELECT training_name FROM training_targets WHERE pmo_category != 'SHE (Safety+Health)'"
                        if filters.get('fiscal_year'):
                            remaining_query += " AND target_year = %s"
                            remaining_params.append(fiscal_year)
                        remaining_query += ")"
                    elif filters['pmo_training_category'] != 'All':
                        remaining_query += " AND t.training_name IN (SELECT training_name FROM training_targets WHERE pmo_category = %s"
                        remaining_params.append(filters['pmo_training_category'])
                        if filters.get('fiscal_year'):
                            remaining_query += " AND target_year = %s"
                            remaining_params.append(fiscal_year)
                        remaining_query += ")"
                
                if filters.get('pl_category') and filters['pl_category'] != 'All':
                    remaining_query += " AND t.training_name IN (SELECT training_name FROM training_targets WHERE pl_category = %s"
                    remaining_params.append(filters['pl_category'])
                    if filters.get('fiscal_year'):
                        remaining_query += " AND target_year = %s"
                        remaining_params.append(fiscal_year)
                    remaining_query += ")"
                
                cursor.execute(remaining_query, remaining_params)
                remaining_result = cursor.fetchone()
                remaining_count = remaining_result['remaining_count'] if remaining_result else 0
                
                # Verify counts add up (for debugging)
                if (matched_count + remaining_count) != tni_metrics['tni_total_count']:
                    print(f"Warning: TNI counts don't match. Total: {tni_metrics['tni_total_count']}, Matched: {matched_count}, Remaining: {remaining_count}")
                
                tni_metrics['matched_count'] = matched_count
                tni_metrics['remaining_count'] = remaining_count
            else:
                tni_metrics['matched_count'] = 0
                tni_metrics['remaining_count'] = 0
            
            # Calculate YTD metrics
            target_month = None
            if filters.get('calendar_month'):
                target_month = filters['calendar_month']
            elif filters.get('month_range_end'):
                target_month = filters['month_range_end']
            
            month_index = get_month_index(target_month)  # Pass selected month
            
            annual_target = target_metrics['target']
            ytd_target = (annual_target // 10) * month_index  # Now uses selected month index
            ytd_actual = total_records
            balance = max(ytd_target - ytd_actual, 0)
            
            percentage_adherence = 0
            if annual_target > 0 and ytd_target > 0:
                raw_adherence = total_records / ytd_target * 100
                # REMOVED: percentage_adherence = round(min(raw_adherence, 100), 1)
                percentage_adherence = round(raw_adherence, 1)  # Allow values above 100%
            
        # FIXED: Get all EOR per_nos to check against with ALL relevant filters
        eor_per_nos = set()
        conn_eor = get_db_connection()
        if conn_eor:
            try:
                query = "SELECT DISTINCT per_no FROM eor_data WHERE 1=1"
                query_params = []
                
                # Apply all relevant filters to EOR data
                factory_filter = filters.get('factory')
                if factory_filter and factory_filter != 'All':
                    query += " AND factory = %s"
                    query_params.append(factory_filter)
                
                # Add gender filter
                gender_filter = filters.get('gender')
                if gender_filter and gender_filter != 'All':
                    query += " AND gender = %s"
                    query_params.append(gender_filter)
                
                # Add employee group filter
                employee_group_filter = filters.get('employee_group')
                if employee_group_filter:
                    query += " AND employee_group = %s"
                    query_params.append(employee_group_filter)
                
                # Add BC No filter
                bc_no_filter = filters.get('bc_no')
                if bc_no_filter:
                    query += " AND bc_no = %s"
                    query_params.append(bc_no_filter)
                
                with conn_eor.cursor() as cursor:
                    cursor.execute(query, query_params)
                    eor_results = cursor.fetchall()
                    eor_per_nos = {row['per_no'] for row in eor_results if row['per_no']}
                    
            except Exception as e:
                print(f"Error fetching EOR data: {str(e)}")
            finally:
                conn_eor.close()
                
        # Get the set of all unique permanent learners in the master_data for the selected filters
        trained_per_nos_any = set()
        # Use a separate connection for this query
        conn_any = get_db_connection()
        if conn_any:
            try:
                query = "SELECT DISTINCT per_no FROM master_data WHERE employee_group = 'PERMANENT' AND per_no IS NOT NULL AND per_no != ''"
                query_params = []
                
                # Apply all relevant filters to match what we did for EOR data
                if factory_filter and factory_filter != 'All':
                    query += " AND factory = %s"
                    query_params.append(factory_filter)
                
                if gender_filter and gender_filter != 'All':
                    query += " AND gender = %s"
                    query_params.append(gender_filter)
                
                if employee_group_filter:
                    query += " AND employee_group = %s"
                    query_params.append(employee_group_filter)
                
                if bc_no_filter:
                    query += " AND bc_no = %s"
                    query_params.append(bc_no_filter)
                
                with conn_any.cursor() as cursor_any:
                    cursor_any.execute(query, query_params)
                    results = cursor_any.fetchall()
                    trained_per_nos_any = {row['per_no'] for row in results}
            except Exception as e:
                print(f"Error fetching unique permanent learners for pending EOR: {str(e)}")
            finally:
                conn_any.close()
                
        # Calculate actual pending EOR (those in EOR but not trained in any training)
        actual_pending_eor = len(eor_per_nos - trained_per_nos_any) if eor_per_nos else 0
        eor_count = len(eor_per_nos) if eor_per_nos else 0
        
        # Calculate hours metrics with pending EOR count
        hours_metrics = calculate_hours_metrics(filters, actual_pending_eor)
        
        return {
            'participant_count': total_records,
            'learning_hours': learning_hours,
            'unique_learners': unique_learners,
            'pending_eor_count': actual_pending_eor,  # Accurate count of EOR employees not trained
            'eor_count': eor_count,  # Total EOR employees
            'total_records': total_records,
            'current_fiscal_year': get_fiscal_year(),
            'target_metrics': target_metrics,
            'tni_metrics': tni_metrics,
            'ytd_metrics': {
                'month_index': month_index,
                'ytd_target': ytd_target,
                'ytd_actual': ytd_actual,
                'balance': balance,
                'annual_target': annual_target,
                'percentage_adherence': percentage_adherence
            },
            'hours_metrics': hours_metrics  # Add the new metrics
            }
            
    except Exception as e:
        print(f"Error calculating dashboard metrics: {str(e)}")
        return None
    finally:
        if conn:
            conn.close()

def get_current_filters(args):
    """Extract current filters from request args"""
    return {
        'per_no': args.get('per_no'),
        'gender': args.get('gender', 'All'),
        'calendar_month': args.get('calendar_month'),
        'month_report_pmo_21_20': args.get('month_report_pmo_21_20'),
        'month_cd_key_26_25': args.get('month_cd_key_26_25'),
        'learning_hours': args.get('learning_hours'),
        'tni_status': args.get('tni_status', ''),  # FIXED: Changed from 'All' to ''
        'training_name': args.get('training_name'),
        'employee_group': args.get('employee_group'),
        'factory': args.get('factory'),
        'start_date': args.get('start_date'),
        'end_date': args.get('end_date'),
        'month_range_start': args.get('month_range_start'),
        'month_range_end': args.get('month_range_end'),
        'pl_category': args.get('pl_category', 'All'),
        'pmo_training_category': args.get('pmo_training_category', 'All'),
        'fiscal_year': args.get('fiscal_year', str(get_fiscal_year()))
    }

def get_employee_hours_breakdown(filters):
    """Get breakdown of SHE and PMO hours for each employee."""
    modified_filters = filters.copy()
    modified_filters['employee_group'] = 'PERMANENT'
    expected_keys = [
        'per_no', 'bc_no', 'gender', 'calendar_month', 'month_report_pmo_21_20',
        'month_cd_key_26_25', 'learning_hours', 'tni_status', 'training_name',
        'employee_group', 'factory', 'start_date', 'end_date', 'month_range_start',
        'month_range_end', 'pl_category', 'pmo_training_category', 'fiscal_year'
    ]
    for key in expected_keys:
        modified_filters.setdefault(key, None)
    base_query, query_params = build_base_query(modified_filters, for_export=True)
    conn = get_db_connection()
    if not conn:
        return {}
    try:
        with conn.cursor() as cursor:
            cursor.execute(base_query, query_params)
            all_records = cursor.fetchall()
            employees = {}
            for record in all_records:
                record_dict = dict(record)
                per_no = record_dict.get('per_no')
                if not per_no:
                    continue
                hours = calculate_learning_hours(record_dict) or 0
                category = record_dict.get('pmo_training_category', '')
                if per_no not in employees:
                    employees[per_no] = {
                        'per_no': per_no,
                        'participants_name': record_dict.get('participants_name', ''),
                        'bc_no': record_dict.get('bc_no', ''),
                        'gender': record_dict.get('gender', ''),
                        'employee_group': record_dict.get('employee_group', ''),
                        'department': record_dict.get('department', ''),
                        'factory': record_dict.get('factory', ''),
                        'she_hours': 0,
                        'pmo_hours': 0,
                        'total_hours': 0
                    }
                if category == 'SHE (Safety+Health)':
                    employees[per_no]['she_hours'] += hours
                else:
                    employees[per_no]['pmo_hours'] += hours
                employees[per_no]['total_hours'] += hours
        
        return employees
    except Exception as e:
        print(f"Error in get_employee_hours_breakdown: {str(e)}")
        return {}
    finally:
        if conn:
            conn.close()

def calculate_hours_metrics(filters, pending_eor_count=0):
    """Calculate metrics for the hours cards"""
    # Get all permanent employees with their SHE and PMO hours
    employees = get_employee_hours_breakdown(filters)
    
    # Calculate metrics
    completed_16 = 0  # 6+ SHE and 10+ PMO
    below_16 = 0
    she_6plus = 0
    she_below_6 = 0
    pmo_10plus = 0
    pmo_below_10 = 0
    cumulative_16plus = 0  # SHE + PMO >= 16
    
    for emp in employees.values():
        # SHE metrics
        if emp['she_hours'] >= 6:
            she_6plus += 1
        else:
            she_below_6 += 1
        
        # PMO metrics
        if emp['pmo_hours'] >= 10:
            pmo_10plus += 1
        else:
            pmo_below_10 += 1
        
        # Combined metrics
        if emp['she_hours'] >= 6 and emp['pmo_hours'] >= 1:
            completed_16 += 1
        else:
            below_16 += 1
            
        # Cumulative metric
        cumulative_hours = emp['she_hours'] + emp['pmo_hours']
        if cumulative_hours >= 16:
            cumulative_16plus += 1
    
    # Add pending EOR counts to the "below" categories
    she_below_6 += pending_eor_count
    pmo_below_10 += pending_eor_count
    below_16 += pending_eor_count
    
    # Update total permanent count to include pending EORs
    total_permanent = len(employees) + pending_eor_count
    
    return {
        'completed_16_count': completed_16,
        'below_16_count': below_16,
        'she_6plus_count': she_6plus,
        'she_below_6_count': she_below_6,
        'pmo_10plus_count': pmo_10plus,
        'pmo_below_10_count': pmo_below_10,
        'cumulative_16plus_count': cumulative_16plus,
        'total_permanent': total_permanent
    }

def get_month_index(month_name=None):
    """Calculate fiscal month index (April=1, May=2, ..., March=12 but capped at 10)"""
    fiscal_month_order = ['April', 'May', 'June', 'July', 'August', 'September', 
                          'October', 'November', 'December', 'January', 'February', 'March']
    
    if month_name:
        try:
            # Find the index in fiscal order
            idx = fiscal_month_order.index(month_name) + 1
            # Cap at 10 for January, February, March
            return min(idx, 10)
        except ValueError:
            # If month not found, fall back to current month
            pass
    
    # If no month provided, use current month
    current_month = datetime.now().month
    if current_month >= 4:  # April to December
        return current_month - 3
    else:  # January to March
        return min(current_month + 9, 10)  # Cap at 10 for January

@view_bp.route('/master_data')
def view_master_data():
    # Get current fiscal year and available years from database
    current_fiscal_year = get_fiscal_year()
    
    # Initialize filters with default values - SET EMPLOYEE_GROUP TO 'Permanent' BY DEFAULT
    filters = {
        'per_no': request.args.get('per_no'),
        'bc_no': request.args.get('bc_no'),
        'gender': request.args.get('gender', 'All'),
        'calendar_month': request.args.get('calendar_month'),
        'month_report_pmo_21_20': request.args.get('month_report_pmo_21_20'),
        'month_cd_key_26_25': request.args.get('month_cd_key_26_25'),
        'learning_hours': request.args.get('learning_hours'),
        'tni_status': request.args.get('tni_status', ''),  # FIXED: Changed from 'All' to ''
        'training_name': request.args.get('training_name'),
        'employee_group': request.args.get('employee_group', 'Permanent'),  # DEFAULT TO 'Permanent'
        'factory': request.args.get('factory'),
        'start_date': request.args.get('start_date'),
        'end_date': request.args.get('end_date'),
        'month_range_start': request.args.get('month_range_start'),
        'month_range_end': request.args.get('month_range_end'),
        'pl_category': request.args.get('pl_category', 'All'),
        'pmo_training_category': request.args.get('pmo_training_category', 'All'),
        'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
    }
    
    # Apply user factory filter based on role
    filters = apply_user_factory_filter(filters)
    
    # Get current page for pagination
    page = request.args.get('page', 1, type=int)
    
    # Calculate dashboard metrics (uses full dataset)
    dashboard_metrics = calculate_dashboard_metrics(filters) or {
    'participant_count': 0,
    'learning_hours': 0,
    'unique_learners': 0,
    'pending_eor_count': 0,
    'eor_count': 0,
    'total_records': 0,
    'current_fiscal_year': current_fiscal_year,
    'target_metrics': {
        'target_hours': 0,
        'target_unique_learners': 0,
        'target': 0
    },
    'tni_metrics': {
        'tni_total_count': 0,
        'tni_unique_learners': 0,
        'matched_count': 0,
        'remaining_count': 0
    },
    'ytd_metrics': {
        'month_index': 0,
        'ytd_target': 0,
        'ytd_actual': 0,
        'balance': 0,
        'annual_target': 0,
        'percentage_adherence': 0
    },
    'hours_metrics': {
        'completed_16_count': 0,
        'below_16_count': 0,
        'she_6plus_count': 0,
        'she_below_6_count': 0,
        'pmo_10plus_count': 0,
        'pmo_below_10_count': 0,
        'cumulative_16plus_count': 0,
        'total_permanent': 0
    }
}
    
    # Calculate total pages
    total_pages = (dashboard_metrics['total_records'] + RECORDS_PER_PAGE - 1) // RECORDS_PER_PAGE
    
    # Get available fiscal years from the database
    conn = get_db_connection()
    fiscal_years = []
    if conn:
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT DISTINCT YEAR(start_date) as year 
                    FROM master_data 
                    WHERE start_date IS NOT NULL
                    ORDER BY year DESC
                """)
                years = [row['year'] for row in cursor.fetchall()]
                
                # Convert to fiscal years (April-March)
                fiscal_years = list({get_fiscal_year(datetime(year, 4, 1)) for year in years})
                fiscal_years.sort(reverse=True)
                
                # If no years found, add current fiscal year
                if not fiscal_years:
                    fiscal_years = [current_fiscal_year]
        except Exception as e:
            print(f"Error fetching fiscal years: {str(e)}")
            fiscal_years = [current_fiscal_year]
        finally:
            conn.close()
    else:
        fiscal_years = [current_fiscal_year]
        
    category_metrics = get_category_metrics(filters)
    monthwise_metrics = get_monthwise_ytd_metrics(filters)
    training_metrics = get_training_wise_metrics(filters)
    annual_metrics = get_annual_ytd_metrics(filters)
    pl_category_counts = get_pl_category_counts(filters)

    # Get employee statistics
    eor_stats = get_employee_group_eor_stats(filters)
    unique_learners_stats = get_unique_learners_permanent(filters)
    
    # Default template variables
    template_vars = {
        'records': [],
        'column_headings': get_column_headings(),
        'eor_stats': eor_stats,  # Add this line
        'unique_learners_stats': unique_learners_stats,
        'title': "Master Data Report",
        'Constants': Constants,
        'filters': filters,
        'category_metrics': category_metrics,
        'monthwise_metrics': monthwise_metrics,
        'pl_category_counts': pl_category_counts,
        'training_metrics': training_metrics,  # Add this line
        'dashboard_metrics': dashboard_metrics,
        'annual_metrics': annual_metrics,  # NEW: Add annual metrics
        'gender_options': ['All', 'Male', 'Female'],
        'all_months': ['January', 'February', 'March', 'April', 'May', 'June', 
                      'July', 'August', 'September', 'October', 'November', 'December'],
        'learning_hours_options': [],
        'tni_options': Constants.TNI_OPTIONS,
        'employee_group_options': ['Permanent','Temporary', 'Contractual', 'Trainee'],
        'training_options': [],
        'learning_hours_values': [2, 4, 6, 8, 16, 24],
        'pl_category_options': ['PL1', 'PL2', 'PL3'],
        'pmo_training_category_options': ['All', 'PMO', 'CESS', 'Digital', 'Functional Skills', 
                                        'Professional Skills', 'SHE (Safety+Health)', 'Sustainability'],
        'current_page': page,
        'records_per_page': RECORDS_PER_PAGE,
        'total_pages': total_pages,
        'total_records': dashboard_metrics['total_records'],
        'fiscal_year_options': fiscal_years,
        'current_fiscal_year': current_fiscal_year,
        # Add timestamp for cache busting
        'timestamp': int(datetime.now().timestamp()),
        'user': get_current_user()  # Add current user info to template
    }
   
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "error")
        return render_template('admin/master_data_table.html', **template_vars)
    try:
        # Get paginated records for display
        base_query, query_params = build_base_query(filters)
        
        with conn.cursor() as cursor:
            cursor.execute(base_query, query_params)
            raw_records = cursor.fetchall()
            records = process_records(raw_records)
            
            # Get unique values for other dropdowns from the database
            cursor.execute("SELECT DISTINCT calendar_month FROM master_data WHERE calendar_month IS NOT NULL ORDER BY calendar_month")
            calendar_month_options = [row['calendar_month'] for row in cursor.fetchall()]
            
            cursor.execute("SELECT DISTINCT month_report_pmo_21_20 FROM master_data WHERE month_report_pmo_21_20 IS NOT NULL ORDER BY month_report_pmo_21_20")
            month_report_pmo_options = [row['month_report_pmo_21_20'] for row in cursor.fetchall()]
            
            cursor.execute("SELECT DISTINCT month_cd_key_26_25 FROM master_data WHERE month_cd_key_26_25 IS NOT NULL ORDER BY month_cd_key_26_25")
            month_cd_key_options = [row['month_cd_key_26_25'] for row in cursor.fetchall()]
            
            cursor.execute("SELECT DISTINCT learning_hours FROM master_data WHERE learning_hours IS NOT NULL ORDER BY learning_hours")
            learning_hours_options = [row['learning_hours'] for row in cursor.fetchall()]
            
            # Get training names from training_names table (all available training programs)
            cursor.execute("SELECT DISTINCT Training_Name FROM training_names ORDER BY Training_Name")
            training_options = [row['Training_Name'] for row in cursor.fetchall()]
            
            template_vars.update({
                'records': records,
                'learning_hours_options': learning_hours_options,
                'calendar_month_options': calendar_month_options,
                'month_report_pmo_options': month_report_pmo_options,
                'month_cd_key_options': month_cd_key_options,
                'training_options': training_options
            })
        return render_template('admin/master_data_table.html', **template_vars)
    except Exception as e:
        flash(f"Error fetching master data: {str(e)}", "error")
        return render_template('admin/master_data_table.html', **template_vars)
    finally:
        if conn:
            conn.close()

# Excel export helper functions
def create_excel_workbook(records, column_headings, title="Report"):
    """Create an Excel workbook with the given records and column headings"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # Write headers with styling
    headers = list(column_headings.values())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_num, record in enumerate(records, 2):  # Start from row 2
        # Write all columns in the specified order
        for col_num, key in enumerate(column_headings.keys(), 1):
            value = str(record.get(key, ''))  # Convert to string to prevent data shifting
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col),
            default=0
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    return wb

@view_bp.route('/download_excel')
def download_excel():
    """Download filtered data as Excel file - exports ALL matching records without pagination"""
    conn = None
    try:
        # Get current fiscal year
        current_fiscal_year = get_fiscal_year()
        
        # Get the same filters as in view_master_data - SET EMPLOYEE_GROUP TO 'Permanent' BY DEFAULT
        filters = {
            'per_no': request.args.get('per_no'),
            'gender': request.args.get('gender', 'All'),
            'calendar_month': request.args.get('calendar_month'),
            'month_report_pmo_21_20': request.args.get('month_report_pmo_21_20'),
            'month_cd_key_26_25': request.args.get('month_cd_key_26_25'),
            'learning_hours': request.args.get('learning_hours'),
            'tni_status': request.args.get('tni_status', ''),  # FIXED: Changed from 'All' to ''
            'training_name': request.args.get('training_name'),
            'employee_group': request.args.get('employee_group', 'Permanent'),  # DEFAULT TO 'Permanent'
            'factory': request.args.get('factory'),
            'start_date': request.args.get('start_date'),
            'end_date': request.args.get('end_date'),
            'month_range_start': request.args.get('month_range_start'),
            'month_range_end': request.args.get('month_range_end'),
            'pl_category': request.args.get('pl_category', 'All'),
            'pmo_training_category': request.args.get('pmo_training_category', 'All'),
            'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
        }
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "error")
            return redirect(url_for('view_bp.view_master_data'))
            
        # Build query without pagination for export
        base_query, query_params = build_base_query(filters, for_export=True)
        
        # Fetch and write data (all records matching filters)
        with conn.cursor() as cursor:
            cursor.execute(base_query, query_params)
            records = cursor.fetchall()
            
            # Process records for export
            processed_records = []
            for row_num, record in enumerate(records, 1):
                record_dict = dict(record)
                
                # Add sequential number starting from 1
                record_dict['sr_no'] = row_num
                
                # Process the record with proper null checks
                day1 = bool(record_dict.get('day_1_attendance', False))
                day2 = bool(record_dict.get('day_2_attendance', False))
                day3 = bool(record_dict.get('day_3_attendance', False))
                
                # Calculate learning hours with fallback to 0 if None
                record_dict['learning_hours'] = calculate_learning_hours(record_dict) or 0
                
                # Format dates and times with null checks
                record_dict['start_date'] = format_date(record_dict.get('start_date')) if record_dict.get('start_date') else ''
                record_dict['end_date'] = format_date(record_dict.get('end_date')) if record_dict.get('end_date') else ''
                record_dict['start_time'] = format_time(record_dict.get('start_time')) if record_dict.get('start_time') else ''
                record_dict['end_time'] = format_time(record_dict.get('end_time')) if record_dict.get('end_time') else ''
                
                # Handle attendance fields
                record_dict['day_1_attendance'] = 'Yes' if day1 else 'No'
                record_dict['day_2_attendance'] = 'Yes' if day2 else 'No'
                record_dict['day_3_attendance'] = 'Yes' if day3 else 'No'
                
                # Clean training name for export with null check
                record_dict['training_name'] = clean_training_name(record_dict.get('training_name', ''))
                
                processed_records.append(record_dict)
        
        # Create workbook
        wb = create_excel_workbook(processed_records, get_column_headings(), "Master Data Report")
        
        # Save workbook to a bytes buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create a filename with timestamp and fiscal year
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"master_data_report_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating Excel report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

def load_eor_data(factory=None):
    """
    Load Employee of Record data from database table with optional factory filter
    
    Args:
        factory: Factory name to filter by. If None or 'All', returns all records
    
    Returns:
        pd.DataFrame: DataFrame containing EOR data
    """
    conn = get_db_connection()
    if not conn:
        return pd.DataFrame()
    
    try:
        query = "SELECT per_no, participants_name, bc_no, gender, employee_group, department, factory FROM eor_data WHERE 1=1"
        query_params = []
        
        # Add factory filter if specified
        if factory and factory != 'All':
            query += " AND factory = %s"
            query_params.append(factory)
            
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            eor_records = cursor.fetchall()
            
            # Convert to DataFrame
            df = pd.DataFrame(eor_records)
            return df
            
    except Exception as e:
        print(f"Error loading EOR data from database: {str(e)}")
        return pd.DataFrame()
    finally:
        if conn:
            conn.close()

def get_eor_count(factory=None):
    """
    Get unique EOR count from database table
    
    Args:
        factory: Factory name to filter by. If None or 'All', returns total count
    
    Returns:
        int: Count of unique employees
    """
    conn = get_db_connection()
    if not conn:
        return 0
    
    try:
        query = "SELECT COUNT(DISTINCT per_no) as eor_count FROM eor_data WHERE 1=1"
        query_params = []
        
        if factory and factory != 'All':
            query += " AND factory = %s"
            query_params.append(factory)
            
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            result = cursor.fetchone()
            return result['eor_count'] if result else 0
            
    except Exception as e:
        print(f"Error getting EOR count: {str(e)}")
        return 0
    finally:
        if conn:
            conn.close()

def get_pending_eor_employees(factory=None):
    """
    Get pending EOR employees (permanent employees in EOR but not in master_data) as a list of dictionaries.
    Args:
        factory: Factory name to filter by. If None or 'All', returns all pending EOR employees.
    Returns:
        list: List of dictionaries representing pending EOR employees.
    """
    # Get EOR data
    eor_df = load_eor_data(factory)
    if eor_df.empty:
        return []
    
    # Get unique learners from master_data (permanent) for the same factory
    conn = get_db_connection()
    if not conn:
        return []
    try:
        unique_query = """
            SELECT DISTINCT per_no FROM master_data
            WHERE employee_group = 'PERMANENT' 
            AND per_no IS NOT NULL
            AND per_no != ''
        """
        query_params = []
        if factory and factory != 'All':
            unique_query += " AND factory = %s"
            query_params.append(factory)
        with conn.cursor() as cursor:
            cursor.execute(unique_query, query_params)
            unique_learners = {row['per_no'] for row in cursor.fetchall()}
        
        # Filter EOR data to get pending records
        pending_eor_df = eor_df[~eor_df['per_no'].isin(unique_learners)]
        
        # Convert to list of dictionaries
        pending_employees = []
        for _, row in pending_eor_df.iterrows():
            emp = {
                'per_no': row['per_no'],
                'participants_name': row.get('participants_name', ''),
                'bc_no': row.get('bc_no', ''),
                'gender': row.get('gender', ''),
                'employee_group': row.get('employee_group', ''),
                'department': row.get('department', ''),
                'factory': row.get('factory', '')
            }
            pending_employees.append(emp)
        return pending_employees
    except Exception as e:
        print(f"Error getting pending EOR employees: {str(e)}")
        return []
    finally:
        conn.close()

@view_bp.route('/download_eor_data')
def download_eor_data():
    """Download Excel of all EOR data with current filters applied"""
    try:
        # Get filters from request
        factory = request.args.get('factory')
        fiscal_year = request.args.get('fiscal_year')
        
        # Apply user factory filter based on role
        filters = {'factory': factory}
        filters = apply_user_factory_filter(filters)
        factory = filters.get('factory')
        
        # Get EOR data using the updated load_eor_data function
        eor_df = load_eor_data(factory)
        
        if eor_df.empty:
            flash("EOR data not available", "error")
            return redirect(url_for('view_bp.view_master_data'))
        
        # Define columns for export
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'PER No'),
            ('participants_name', 'Name'),
            ('bc_no', 'BC No'),
            ('gender', 'Gender'),
            ('employee_group', 'Employee Group'),
            ('department', 'Department'),
            ('factory', 'Factory')
        ]
        
        # Process records for export (add serial numbers)
        processed_records = []
        for row_num, (_, row) in enumerate(eor_df.iterrows(), 1):
            record_dict = dict(row)
            record_dict['sr_no'] = row_num
            processed_records.append(record_dict)
        
        # Create workbook with styled headers
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, "EOR Data")
        
        # Apply gold color to headers (consistent with other exports)
        ws = wb.active
        for cell in ws[1]:  # First row
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"eor_data_FY{fiscal_year}_{timestamp}.xlsx" if fiscal_year else f"eor_data_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating EOR data report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))

@view_bp.route('/download_pending_eor')
def download_pending_eor():
    """Download Excel of pending EOR (EOR count - unique learners)"""
    conn = None
    try:
        # Get current fiscal year
        current_fiscal_year = get_fiscal_year()
        
        # Get filters
        filters = {
            'factory': request.args.get('factory'),
            'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
        }
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        # Get EOR data from database
        conn_eor = get_db_connection()
        if not conn_eor:
            flash("Database connection failed", "error")
            return redirect(url_for('view_bp.view_master_data'))
        try:
            query = "SELECT per_no, participants_name, bc_no, gender, employee_group, department, factory FROM eor_data WHERE 1=1"
            query_params = []
            
            factory_filter = filters.get('factory')
            if factory_filter and factory_filter != 'All':
                query += " AND factory = %s"
                query_params.append(factory_filter)
                
            with conn_eor.cursor() as cursor:
                cursor.execute(query, query_params)
                eor_records = cursor.fetchall()
                
            if not eor_records:
                flash("EOR data not available", "error")
                return redirect(url_for('view_bp.view_master_data'))
                
            # Convert to DataFrame for consistency with existing code
            eor_df = pd.DataFrame(eor_records)
            
        finally:
            conn_eor.close()
        
        # Get unique learners from master_data
        unique_query = """
            SELECT DISTINCT per_no, participants_name, bc_no, gender, 
                   employee_group, department, factory
            FROM master_data
            WHERE employee_group = 'PERMANENT' 
            AND per_no IS NOT NULL
            AND per_no != ''
        """
        unique_params = []
        
        # Apply factory filter to unique learners query
        if factory_filter and factory_filter != 'All':
            unique_query += " AND factory = %s"
            unique_params.append(factory_filter)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "error")
            return redirect(url_for('view_bp.view_master_data'))
        
        with conn.cursor() as cursor:
            cursor.execute(unique_query, unique_params)
            unique_learners = {row['per_no'] for row in cursor.fetchall()}
        
        # Filter EOR data to get pending records (not in unique learners)
        pending_eor_df = eor_df[~eor_df['per_no'].isin(unique_learners)]
        
        # Define columns
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'PER No'),
            ('participants_name', 'Name'),
            ('bc_no', 'BC No'),
            ('gender', 'Gender'),
            ('employee_group', 'Employee Group'),
            ('department', 'Department'),
            ('factory', 'Factory')
        ]
        
        # Process records for export
        processed_records = []
        for row_num, (_, row) in enumerate(pending_eor_df.iterrows(), 1):
            record_dict = dict(row)
            record_dict['sr_no'] = row_num
            processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, "Pending EOR")
        
        # Apply gold color to headers
        ws = wb.active
        for cell in ws[1]:  # First row
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pending_eor_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating pending EOR report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()
            
# Generic download functions to reduce code duplication
def download_filtered_hours_report(min_hours=None, max_hours=None, category_filter=None,
                                   exclude_she=False, title=None, filename_prefix=None,
                                   include_pending_eor=False):
    """Download filtered hours reports using same aggregation logic as metrics (grouped per unique per_no)."""
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "error")
        return redirect(url_for('view_bp.view_master_data'))
    try:
        # Get current filters (fiscal year, etc.)
        filters = get_current_filters(request.args)
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        # Get aggregated employee hours (same logic as metrics)
        employees = get_employee_hours_breakdown(filters)
        filtered_records = []
        for emp in employees.values():
            # Determine hours to compare depending on category filter
            if category_filter == "SHE (Safety+Health)":
                hours = emp.get("she_hours", 0)
            elif category_filter == "PMO":
                hours = emp.get("pmo_hours", 0)
            else:
                # If no category filter, combine both
                hours = emp.get("she_hours", 0) + emp.get("pmo_hours", 0)
            # Apply permanent employee filter
            if emp.get("employee_group") != "Permanent":
                continue
            # Apply hour filters
            if min_hours is not None and hours < min_hours:
                continue
            if max_hours is not None and hours >= max_hours:
                continue
            filtered_records.append({
                'per_no': emp['per_no'],
                'participants_name': emp.get('participants_name', ''),
                'bc_no': emp.get('bc_no', ''),
                'gender': emp.get('gender', ''),
                'employee_group': emp.get('employee_group', ''),
                'department': emp.get('department', ''),
                'factory': emp.get('factory', ''),
                'learning_hours': hours
            })
        
        # Include pending EOR employees if requested
        if include_pending_eor:
            factory = filters.get('factory')
            pending_employees = get_pending_eor_employees(factory=factory)
            for emp in pending_employees:
                record = {
                    'per_no': emp['per_no'],
                    'participants_name': emp.get('participants_name', ''),
                    'bc_no': emp.get('bc_no', ''),
                    'gender': emp.get('gender', ''),
                    'employee_group': emp.get('employee_group', ''),
                    'department': emp.get('department', ''),
                    'factory': emp.get('factory', ''),
                    'learning_hours': 0   # because they have no training
                }
                filtered_records.append(record)
        
        # Define columns
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'Per. No'),
            ('participants_name', 'Participants Name'),
            ('bc_no', 'BC No'),
            ('gender', 'Gender'),
            ('employee_group', 'Employee Group'),
            ('department', 'Department'),
            ('factory', 'Factory'),
            ('learning_hours', 'Learning Hours')
        ]
        
        # Process records for export
        processed_records = []
        for idx, emp in enumerate(filtered_records, 1):
            record_dict = emp.copy()
            record_dict['sr_no'] = idx
            processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, title or "Hours Report")
        
        # Define color fills
        fill_zero = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light red
        fill_sixteen = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Light green
        
        # Apply color to entire rows if needed
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            learning_hours = int(row[8].value)  # Learning hours is in column 9 (index 8)
            if learning_hours == 0:
                for cell in row:
                    cell.fill = fill_zero
            elif learning_hours == 10:
                for cell in row:
                    cell.fill = fill_sixteen
        
        # Return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error generating {title}: {str(e)}")
        flash(f"Error generating report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

def download_combined_hours_report(she_min_hours=6, pmo_min_hours=10, incomplete_only=False,
                                 title=None, filename_prefix=None,
                                 include_pending_eor=False):
    """Download Excel of employees with combined SHE and PMO hours"""
    conn = None
    try:
        # Get current filters
        filters = get_current_filters(request.args)
        current_fiscal_year = get_fiscal_year()
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        # Get all permanent employees with their SHE and PMO hours
        employees = get_employee_hours_breakdown(filters)
        
        if not employees and not include_pending_eor:
            flash("No data found matching the criteria.", "warning")
            return redirect(url_for('view_bp.view_master_data'))
        
        # Filter employees based on criteria
        filtered_employees = []
        for emp in employees.values():
            she_ok = emp['she_hours'] >= she_min_hours
            pmo_ok = emp['pmo_hours'] >= pmo_min_hours
            
            if incomplete_only:
                # Include if they DON'T meet both criteria
                if not (she_ok and pmo_ok):
                    filtered_employees.append(emp)
            else:
                # Include if they meet BOTH criteria
                if she_ok and pmo_ok:
                    filtered_employees.append(emp)
        
        # Include pending EOR employees if requested
        if include_pending_eor:
            factory = filters.get('factory')
            pending_employees = get_pending_eor_employees(factory=factory)
            for emp in pending_employees:
                # Create a record for pending EOR employee with 0 hours
                record = {
                    'per_no': emp['per_no'],
                    'participants_name': emp.get('participants_name', ''),
                    'bc_no': emp.get('bc_no', ''),
                    'gender': emp.get('gender', ''),
                    'employee_group': emp.get('employee_group', ''),
                    'department': emp.get('department', ''),
                    'factory': emp.get('factory', ''),
                    'she_hours': 0,
                    'pmo_hours': 0,
                    'total_hours': 0
                }
                filtered_employees.append(record)
        
        if not filtered_employees:
            flash("No employees found matching the criteria.", "warning")
            return redirect(url_for('view_bp.view_master_data'))
        
        # Define columns
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'Per. No'),
            ('participants_name', 'Participants Name'),
            ('bc_no', 'BC No'),
            ('gender', 'Gender'),
            ('employee_group', 'Employee Group'),
            ('department', 'Department'),
            ('factory', 'Factory'),
            ('she_hours', 'SHE Hours'),
            ('pmo_hours', 'PMO Hours'),
            ('total_hours', 'Total Hours')
        ]
        
        # Process records for export
        processed_records = []
        for row_num, emp in enumerate(filtered_employees, 1):
            record_dict = emp.copy()
            record_dict['sr_no'] = row_num 
            processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, title or "Combined Hours Report")
        
        # Apply color coding based on completion status
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if incomplete_only:
                # Highlight incomplete records in light red
                for cell in row:
                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            else:
                # Highlight complete records in light green
                for cell in row:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error generating {title}: {str(e)}")  # Add debugging
        flash(f"Error generating report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

def download_cumulative_hours_report(min_hours=None, title=None, filename_prefix=None):
    """Download cumulative hours reports (SHE + PMO) - Only for 16+ hours"""
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "error")
        return redirect(url_for('view_bp.view_master_data'))
    try:
        # Get current filters
        filters = get_current_filters(request.args)
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        # Get aggregated employee hours
        employees = get_employee_hours_breakdown(filters)
        filtered_records = []
        for emp in employees.values():
            # Calculate cumulative hours (SHE + PMO)
            cumulative_hours = emp.get("she_hours", 0) + emp.get("pmo_hours", 0)
            
            # Apply permanent employee filter
            if emp.get("employee_group") != "Permanent":
                continue
            # Apply hour filter - only include 16+ hours
            if min_hours is not None and cumulative_hours < min_hours:
                continue
            filtered_records.append({
                'per_no': emp['per_no'],
                'participants_name': emp.get('participants_name', ''),
                'bc_no': emp.get('bc_no', ''),
                'gender': emp.get('gender', ''),
                'employee_group': emp.get('employee_group', ''),
                'department': emp.get('department', ''),
                'factory': emp.get('factory', ''),
                'she_hours': emp.get('she_hours', 0),
                'pmo_hours': emp.get('pmo_hours', 0),
                'cumulative_hours': cumulative_hours
            })
        
        # Define columns
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'Per. No'),
            ('participants_name', 'Participants Name'),
            ('bc_no', 'BC No'),
            ('gender', 'Gender'),
            ('employee_group', 'Employee Group'),
            ('department', 'Department'),
            ('factory', 'Factory'), 
            ('she_hours', 'SHE Hours'),
            ('pmo_hours', 'PMO Hours'),
            ('cumulative_hours', 'Cumulative Hours')
        ]
        
        # Process records for export
        processed_records = []
        for idx, emp in enumerate(filtered_records, 1):
            record_dict = emp.copy()
            record_dict['sr_no'] = idx
            processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, title or "Cumulative Hours Report")
        
        # Define color fill for 16+ hours (green)
        fill_16plus = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Light green
        
        # Apply green color to all rows (all are 16+ hours)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = fill_16plus
        
        # Return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error generating {title}: {str(e)}")
        flash(f"Error generating report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

# Specific download routes using the generic functions
@view_bp.route('/download_she_6plus_hours')
def download_she_6plus_hours():
    """Download Excel of SHE trainings with 6+ hours"""
    return download_filtered_hours_report(
        min_hours=6,
        category_filter="SHE (Safety+Health)",
        title="SHE 6+ Hours Report",
        filename_prefix="she_6plus_hours"
    )

@view_bp.route('/download_she_below_6_hours')
def download_she_below_6_hours():
    """Download Excel of SHE trainings below 6 hours"""
    return download_filtered_hours_report(
        max_hours=6,
        category_filter="SHE (Safety+Health)",
        title="SHE Below 6 Hours Report",
        filename_prefix="she_below_6_hours",
        include_pending_eor=True
    )

@view_bp.route('/download_pmo_10plus_hours')
def download_pmo_10plus_hours():
    """Download Excel of PMO trainings (excluding SHE) with 10+ hours"""
    return download_filtered_hours_report(
        min_hours=10,
        category_filter="PMO",
        exclude_she=True,
        title="PMO 10+ Hours Report",
        filename_prefix="pmo_10plus_hours"
    )

@view_bp.route('/download_pmo_below_10_hours')
def download_pmo_below_10_hours():
    """Download Excel of PMO trainings (excluding SHE) below 10 hours"""
    return download_filtered_hours_report(
        max_hours=10,
        category_filter="PMO",
        exclude_she=True,
        title="PMO Below 10 Hours Report",
        filename_prefix="pmo_below_10_hours",
        include_pending_eor=True
    )

@view_bp.route('/download_completed_16_hours')
def download_completed_16_hours():
    """Download Excel of employees who completed 16 hours (6+ SHE and 10+ PMO)"""
    return download_combined_hours_report(
        she_min_hours=6,
        pmo_min_hours=1,
        title="Completed 16 Hours Report",
        filename_prefix="completed_16_hours"
    )

@view_bp.route('/download_incomplete_16_hours')
def download_incomplete_16_hours():
    """Download Excel of employees who didn't complete 16 hours"""
    return download_combined_hours_report(
        she_min_hours=6,
        pmo_min_hours=10,
        incomplete_only=True,
        title="Incomplete 16 Hours Report",
        filename_prefix="incomplete_16_hours",
        include_pending_eor=True
    )

@view_bp.route('/download_cumulative_16plus_hours')
def download_cumulative_16plus_hours():
    """Download Excel of employees with cumulative 16+ hours (SHE + PMO)"""
    return download_cumulative_hours_report(
        min_hours=16,
        title="Cumulative 16+ Hours Report",
        filename_prefix="cumulative_16plus_hours"
    )

@view_bp.route('/download_unique_learners')
def download_unique_learners():
    """Download unique learners data as Excel file with proper handling of missing PER NO"""
    conn = None
    try:
        # Get current fiscal year
        current_fiscal_year = get_fiscal_year()
        
        # Get filters
        filters = {
            'per_no': request.args.get('per_no'),
            'gender': request.args.get('gender', 'All'),
            'calendar_month': request.args.get('calendar_month'),
            'month_report_pmo_21_20': request.args.get('month_report_pmo_21_20'),
            'month_cd_key_26_25': request.args.get('month_cd_key_26_25'),
            'learning_hours': request.args.get('learning_hours'),
            'tni_status': request.args.get('tni_status', ''),  # FIXED: Changed from 'All' to ''
            'training_name': request.args.get('training_name'),
            'employee_group': request.args.get('employee_group'),
            'factory': request.args.get('factory'),
            'start_date': request.args.get('start_date'),
            'end_date': request.args.get('end_date'),
            'month_range_start': request.args.get('month_range_start'),
            'month_range_end': request.args.get('month_range_end'),
            'pl_category': request.args.get('pl_category', 'All'),
            'pmo_training_category': request.args.get('pmo_training_category', 'All'),
            'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
        }
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "error")
            return redirect(url_for('view_bp.view_master_data'))
            
        # Build query to get unique learners
        base_query = """
            SELECT 
                per_no,
                participants_name,
                bc_no,
                gender,
                employee_group,
                department,
                factory
            FROM master_data
            WHERE employee_group = 'PERMANENT'
        """
        query_params = []
        
        # Apply all the same filters as in build_base_query
        base_query, query_params = apply_standard_filters(base_query, query_params, filters)
        
        # Group by to get unique learners
        base_query += " GROUP BY per_no, participants_name, bc_no, gender, employee_group, department, factory"
        
        # Define columns with proper order
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'Per. No'),
            ('participants_name', 'Participants Name'),
            ('bc_no', 'BC No'),
            ('gender', 'Gender'),
            ('employee_group', 'Employee group'),
            ('department', 'Department'), 
            ('factory', 'Factory')
        ]
        
        # Fetch and write data
        with conn.cursor() as cursor:
            cursor.execute(base_query, query_params)
            records = cursor.fetchall()
            
            # Process records for export
            processed_records = []
            for row_num, record in enumerate(records, 1):
                record_dict = dict(record)
                record_dict['sr_no'] = row_num
                
                # Ensure all columns exist with empty string if missing
                for key, header in columns:
                    if key not in record_dict:
                        record_dict[key] = ''
                
                processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, "Unique Learners Report")
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"unique_learners_report_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

@view_bp.route('/download_tni_shared')
def download_tni_shared():
    """Download Excel of all TNI shared data"""
    conn = None
    try:
        # Get current fiscal year
        current_fiscal_year = get_fiscal_year()
        
        # Get filters
        filters = {
            'factory': request.args.get('factory'),
            'training_name': request.args.get('training_name'),
            'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
        }
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "error")
            return redirect(url_for('view_bp.view_master_data'))
            
        # Query to get all TNI shared data
        query = """
            SELECT 
                per_no,
                name,
                factory,
                training_name,
                hours
            FROM tni_data
            WHERE 1=1
        """
        query_params = []
        
        # Apply fiscal year filter (removed since we don't have date fields)
        
        # Apply other filters
        if filters['factory']:
            query += " AND factory = %s"
            query_params.append(filters['factory'])
        
        if filters['training_name']:
            query += " AND training_name = %s"
            query_params.append(filters['training_name'])
        
        # Define columns based on available fields
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'PER No'),
            ('name', 'Name'),
            ('factory', 'Factory'),
            ('training_name', 'Training Name'),
            ('hours', 'Training Hours')
        ]
        
        # Fetch and write data
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            records = cursor.fetchall()
            
            # Process records for export
            processed_records = []
            for row_num, record in enumerate(records, 1):
                record_dict = dict(record)
                record_dict['sr_no'] = row_num
                processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, "TNI Shared Data")
        
        # Freeze header row
        ws = wb.active
        ws.freeze_panes = 'A2'
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TNI_shared_data_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating TNI shared report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

@view_bp.route('/download_tni_matched')
def download_tni_matched():
    """Download Excel of participants who matched TNI plan (attended as planned)"""
    conn = None
    try:
        # Get current fiscal year
        current_fiscal_year = get_fiscal_year()
        
        # Get filters
        filters = {
            'factory': request.args.get('factory'),
            'training_name': request.args.get('training_name'),
            'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
        }
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "error")
            return redirect(url_for('view_bp.view_master_data'))
            
        # Query to find TNI records that have matching master_data records
        query = """
            SELECT 
                t.per_no,
                t.name,
                t.factory,
                t.training_name,
                t.hours as planned_hours,
                m.learning_hours as actual_hours,
                m.start_date,
                m.end_date,
                m.day_1_attendance,
                m.day_2_attendance,
                m.day_3_attendance
            FROM tni_data t
            JOIN master_data m ON t.per_no = m.per_no 
                AND t.factory = m.factory
                AND t.training_name = m.training_name
            WHERE 1=1
        """
        query_params = []
        
        # Apply other filters
        if filters['factory']:
            query += " AND t.factory = %s"
            query_params.append(filters['factory'])
        
        if filters['training_name']:
            query += " AND t.training_name = %s"
            query_params.append(filters['training_name'])
        
        # Define columns with available data
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'PER No'),
            ('name', 'Name'),
            ('factory', 'Factory'),
            ('training_name', 'Training Name'),
            ('planned_hours', 'Planned Hours'),
            ('actual_hours', 'Actual Hours'),
            ('start_date', 'Start Date'),
            ('end_date', 'End Date'),
            ('day_1_attendance', 'Day 1 Attendance'),
            ('day_2_attendance', 'Day 2 Attendance'),
            ('day_3_attendance', 'Day 3 Attendance')
        ]
        
        # Fetch and write data
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            records = cursor.fetchall()
            
            # Process records for export
            processed_records = []
            for row_num, record in enumerate(records, 1):
                record_dict = dict(record)
                record_dict['sr_no'] = row_num
                
                # Format dates
                record_dict['start_date'] = format_date(record_dict.get('start_date'))
                record_dict['end_date'] = format_date(record_dict.get('end_date'))
                
                processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, "TNI Matched Participants")
        
        # Apply highlighting for attendance
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=12):
            for cell in row:
                if cell.value.lower() == 'yes':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TNI_matched_participants_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating matched TNI report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

@view_bp.route('/download_tni_remaining')
def download_tni_remaining():
    """Download Excel of participants in TNI plan who haven't attended yet"""
    conn = None
    try:
        # Get current fiscal year
        current_fiscal_year = get_fiscal_year()
        
        # Get filters
        filters = {
            'factory': request.args.get('factory'),
            'training_name': request.args.get('training_name'),
            'fiscal_year': request.args.get('fiscal_year', str(current_fiscal_year))
        }
        
        # Apply user factory filter based on role
        filters = apply_user_factory_filter(filters)
        
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "error")
            return redirect(url_for('view_bp.view_master_data'))
            
        # Query to find TNI records without matching master_data records
        query = """
            SELECT 
                t.per_no,
                t.name,
                t.factory,
                t.training_name,
                t.hours
            FROM tni_data t
            LEFT JOIN master_data m ON t.per_no = m.per_no 
                AND t.factory = m.factory
                AND t.training_name = m.training_name
            WHERE m.per_no IS NULL
        """
        query_params = []
        
        # Apply other filters
        if filters['factory']:
            query += " AND t.factory = %s"
            query_params.append(filters['factory'])
        
        if filters['training_name']:
            query += " AND t.training_name = %s"
            query_params.append(filters['training_name'])
        
        # Define columns with available data
        columns = [
            ('sr_no', 'SR.No'),
            ('per_no', 'PER No'),
            ('name', 'Name'),
            ('factory', 'Factory'),
            ('training_name', 'Training Name'),
            ('hours', 'Planned Hours')
        ]
        
        # Fetch and write data
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            records = cursor.fetchall()
            
            # Process records for export
            processed_records = []
            for row_num, record in enumerate(records, 1):
                record_dict = dict(record)
                record_dict['sr_no'] = row_num
                processed_records.append(record_dict)
        
        # Create workbook
        column_headings = {key: header for key, header in columns}
        wb = create_excel_workbook(processed_records, column_headings, "TNI Remaining Participants")
        
        # Freeze header row
        ws = wb.active
        ws.freeze_panes = 'A2'
        
        # Create and return file
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TNI_remaining_participants_FY{filters['fiscal_year']}_{timestamp}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating remaining TNI report: {str(e)}", "error")
        return redirect(url_for('view_bp.view_master_data'))
    finally:
        if conn:
            conn.close()

def get_pl_category_counts(filters):
    """Get counts of unique permanent learners by PL category with annual targets and YTD metrics"""
    conn = get_db_connection()
    if not conn:
        return {
            'PL1': {
                'unique_learners_count': 0,
                'annual_target': 0,
                'ytd_coverage': 0,
                'ytd_target': 0,
                'adherence': 0
            },
            'PL2': {
                'unique_learners_count': 0,
                'annual_target': 0,
                'ytd_coverage': 0,
                'ytd_target': 0,
                'adherence': 0
            },
            'PL3': {
                'unique_learners_count': 0,
                'annual_target': 0,
                'ytd_coverage': 0,
                'ytd_target': 0,
                'adherence': 0
            }
        }
    
    try:
        # Build query to get counts by PL category for permanent employees
        query = """
            SELECT 
                pl_category,
                COUNT(DISTINCT per_no) as count
            FROM master_data
            WHERE employee_group = 'PERMANENT'
            AND per_no IS NOT NULL
            AND per_no != ''
        """
        query_params = []
        
        # Apply all filters
        query, query_params = apply_standard_filters(query, query_params, filters)
        
        # Group by PL category
        query += " GROUP BY pl_category"
        
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            results = cursor.fetchall()
            
            # Initialize counts
            counts = {'PL1': 0, 'PL2': 0, 'PL3': 0}
            
            # Update counts from results
            for row in results:
                pl_category = row['pl_category']
                if pl_category in counts:
                    counts[pl_category] = row['count']
            
            # Get current month index
            month_index = get_month_index()
            
            # Initialize result dictionary with additional metrics
            pl_metrics = {}
            
            # For each PL category, get annual target and YTD metrics
            for category in ['PL1', 'PL2', 'PL3']:
                # Create category-specific filters
                category_filters = filters.copy()
                category_filters['pl_category'] = category
                
                # Get dashboard metrics for this category
                dashboard_metrics = calculate_dashboard_metrics(category_filters)
                
                if dashboard_metrics:
                    annual_target = dashboard_metrics['target_metrics']['target']
                    ytd_actual = dashboard_metrics['ytd_metrics']['ytd_actual']
                    
                    # Calculate YTD target
                    if annual_target > 0 and month_index > 0:
                        ytd_target = (annual_target / 10) * month_index
                    else:
                        ytd_target = 0
                    
                    # Calculate adherence percentage
                    adherence = 0
                    if ytd_target > 0:
                        # REMOVED: adherence = min((ytd_actual / ytd_target) * 100, 100)
                        adherence = (ytd_actual / ytd_target) * 100  # Allow values above 100%
                    
                    pl_metrics[category] = {
                        'unique_learners_count': counts[category],
                        'annual_target': annual_target,
                        'ytd_coverage': ytd_actual,
                        'ytd_target': int(ytd_target),
                        'adherence': round(adherence, 1)
                    }
                else:
                    # If no metrics available, use zeros
                    pl_metrics[category] = {
                        'unique_learners_count': counts[category],
                        'annual_target': 0,
                        'ytd_coverage': 0,
                        'ytd_target': 0,
                        'adherence': 0
                    }
            
            return pl_metrics
            
    except Exception as e:
        print(f"Error getting PL category counts: {str(e)}")
        return {
            'PL1': {
                'unique_learners_count': 0,
                'annual_target': 0,
                'ytd_coverage': 0,
                'ytd_target': 0,
                'adherence': 0
            },
            'PL2': {
                'unique_learners_count': 0,
                'annual_target': 0,
                'ytd_coverage': 0,
                'ytd_target': 0,
                'adherence': 0
            },
            'PL3': {
                'unique_learners_count': 0,
                'annual_target': 0,
                'ytd_coverage': 0,
                'ytd_target': 0,
                'adherence': 0
            }
        }
    finally:
        if conn:
            conn.close()

from datetime import datetime

def get_category_metrics(filters):
    """Calculate category-wise metrics for training categories"""
    categories = ['PMO', 'CESS', 'Digital', 'Functional Skills', 
                 'Professional Skills', 'SHE (Safety+Health)', 'Sustainability']
    
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            return []
        
        # Get current fiscal month number (1=April, 12=March)
        now = datetime.now()
        fiscal_year_start_month = 4  # April
        fiscal_month_index = (now.month - fiscal_year_start_month + 12) % 12 + 1
        
        # Get Permanent EOR count
        eor_per_nos = set()
        query = "SELECT DISTINCT per_no FROM eor_data WHERE employee_group = 'Permanent'"
        query_params = []
        
        factory_filter = filters.get('factory')
        if factory_filter and factory_filter != 'All':
            query += " AND factory = %s"
            query_params.append(factory_filter)
            
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            eor_results = cursor.fetchall()
            eor_per_nos = {row['per_no'] for row in eor_results if row['per_no']}
        
        permanent_eor_count = len(eor_per_nos)
        # ✅ Calculate EOR YTD Target
        eor_ytd_target = int((permanent_eor_count / 10) * fiscal_month_index)
        
        # ---- Rest of your existing code ----
        trained_per_nos_any = set()
        query = "SELECT DISTINCT per_no FROM master_data WHERE employee_group = 'PERMANENT' AND per_no IS NOT NULL AND per_no != ''"
        query_params = []
        if factory_filter and factory_filter != 'All':
            query += " AND factory = %s"
            query_params.append(factory_filter)
            
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            results = cursor.fetchall()
            trained_per_nos_any = {row['per_no'] for row in results}
        
        actual_pending_eor = len(eor_per_nos - trained_per_nos_any) if eor_per_nos else 0
        all_employees = get_employee_hours_breakdown(filters)
        
        completed_16_employees = set()
        for per_no, emp in all_employees.items():
            if emp['she_hours'] >= 6 and emp['pmo_hours'] >= 1:
                completed_16_employees.add(per_no)
        
        category_employees_map = {}
        for category in categories:
            category_filters = filters.copy()
            category_filters['pmo_training_category'] = category
            
            query = """
                SELECT DISTINCT per_no 
                FROM master_data 
                WHERE employee_group = 'PERMANENT' 
                AND per_no IS NOT NULL 
                AND per_no != ''
            """
            query_params = []
            query, query_params = apply_standard_filters(query, query_params, category_filters)
            
            with conn.cursor() as cursor:
                cursor.execute(query, query_params)
                emp_results = cursor.fetchall()
                category_employees_map[category] = {row['per_no'] for row in emp_results}
        
        results = []
        
        for category in categories:
            category_filters = filters.copy()
            category_filters['pmo_training_category'] = category
            
            metrics = calculate_dashboard_metrics(category_filters)
            hours_metrics = calculate_hours_metrics(category_filters, 0)
            pl_counts = get_pl_category_counts(category_filters)
            
            completed_16_in_category = len(completed_16_employees & category_employees_map[category])
            
            if metrics and hours_metrics:
                annual_target = metrics['target_metrics']['target']
                ytd_target = metrics['ytd_metrics']['ytd_target']
                ytd_actual = metrics['ytd_metrics']['ytd_actual']
                
                adherence = (ytd_actual / ytd_target * 100) if ytd_target > 0 else 0
                tni_total_count = metrics['tni_metrics']['tni_total_count']
                tni_matched_count = metrics['tni_metrics']['matched_count']
                tni_adherence = (tni_matched_count / tni_total_count * 100) if tni_total_count > 0 else 0
                unique_learners = metrics['unique_learners']
                ul_adherence = (unique_learners / permanent_eor_count * 100) if permanent_eor_count > 0 else 0
                
                results.append({
                    'category': category,
                    'annual_target': annual_target,
                    'ytd_target': ytd_target,
                    'ytd_actual': ytd_actual,
                    'adherence': round(adherence, 1),
                    'unique_learners': unique_learners,
                    'learning_hours': metrics['learning_hours'],
                    'eor_count': permanent_eor_count,
                    'pending_eor_count': actual_pending_eor,
                    'permanent_eor_count': permanent_eor_count,
                    'eor_ytd_target': eor_ytd_target,  # ✅ ADD THIS
                    'tni_total_count': tni_total_count,
                    'tni_matched_count': tni_matched_count,
                    'tni_adherence': round(tni_adherence, 1),
                    'ul_adherence': round(ul_adherence, 1),
                    'pl1_count': pl_counts.get('PL1', 0),
                    'pl2_count': pl_counts.get('PL2', 0),
                    'pl3_count': pl_counts.get('PL3', 0),
                    'she_6plus_count': hours_metrics.get('she_6plus_count', 0),
                    'she_below_6_count': hours_metrics.get('she_below_6_count', 0),
                    'pmo_10plus_count': hours_metrics.get('pmo_10plus_count', 0),
                    'pmo_below_10_count': hours_metrics.get('pmo_below_10_count', 0),
                    'completed_16_count': completed_16_in_category,
                    'cumulative_16plus_count': hours_metrics.get('cumulative_16plus_count', 0),
                    'below_16_count': hours_metrics.get('below_16_count', 0),
                    'total_permanent': hours_metrics.get('total_permanent', 0)
                })
            else:
                results.append({
                    'category': category,
                    'annual_target': 0,
                    'ytd_target': 0,
                    'ytd_actual': 0,
                    'adherence': 0,
                    'unique_learners': 0,
                    'learning_hours': 0,
                    'eor_count': permanent_eor_count,
                    'pending_eor_count': actual_pending_eor,
                    'permanent_eor_count': permanent_eor_count,
                    'eor_ytd_target': eor_ytd_target,  # ✅ ADD THIS HERE TOO
                    'tni_total_count': 0,
                    'tni_matched_count': 0,
                    'tni_adherence': 0,
                    'ul_adherence': 0,
                    'pl1_count': 0,
                    'pl2_count': 0,
                    'pl3_count': 0,
                    'she_6plus_count': 0,
                    'she_below_6_count': 0,
                    'pmo_10plus_count': 0,
                    'pmo_below_10_count': 0,
                    'completed_16_count': 0,
                    'cumulative_16plus_count': 0,
                    'below_16_count': 0,
                    'total_permanent': 0
                })
        
        return results
        
    except Exception as e:
        print(f"Error calculating category metrics: {str(e)}")
        return []
    finally:
        if conn:
            conn.close()


def get_monthwise_ytd_metrics(filters):
    """Calculate month-wise YTD target and coverage for the fiscal year"""
    # Create a copy of filters and remove month-specific filters for the chart
    chart_filters = filters.copy()
    chart_filters.pop('calendar_month', None)
    chart_filters.pop('month_range_start', None)
    chart_filters.pop('month_range_end', None)
    chart_filters.pop('month_report_pmo_21_20', None)
    chart_filters.pop('month_cd_key_26_25', None)
    
    # Get the annual target from dashboard metrics using modified filters
    dashboard_metrics = calculate_dashboard_metrics(chart_filters)
    annual_target = dashboard_metrics['target_metrics']['target'] if dashboard_metrics else 0
    
    # Define fiscal year month order (April to March)
    fiscal_month_order = ['April', 'May', 'June', 'July', 'August', 'September', 
                          'October', 'November', 'December', 'January', 'February', 'March']
    
    # Initialize monthly counts
    monthly_counts = {month: 0 for month in fiscal_month_order}
    
    # Get monthly actual counts from database
    conn = get_db_connection()
    if not conn:
        return []
    
    try:
        # Build query to get monthly counts
        query = """
            SELECT calendar_month, COUNT(id) as count
            FROM master_data
            WHERE 1=1
        """
        query_params = []
        
        # Apply fiscal year filter
        query, query_params = apply_fiscal_year_filter(query, query_params, chart_filters.get('fiscal_year'))
        
        # Apply other filters (using chart_filters which excludes month filters)
        query, query_params = apply_standard_filters(query, query_params, chart_filters)
        
        # Group by calendar month
        query += " GROUP BY calendar_month"
        
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            results = cursor.fetchall()
            
            # Update monthly counts from results
            for row in results:
                month = row['calendar_month']
                if month in monthly_counts:
                    monthly_counts[month] = row['count']
        
        # Calculate cumulative metrics for each month
        results = []
        cumulative_coverage = 0
        
        for i, month in enumerate(fiscal_month_order, 1):
            # Get count for this month
            monthly_count = monthly_counts[month]
            cumulative_coverage += monthly_count
            
            # Calculate month index (capped at 10)
            month_index = i
            capped_index = min(month_index, 10)
            
            # Calculate YTD target for this month
            if annual_target > 0:
                ytd_target = (annual_target / 10) * capped_index
            else:
                ytd_target = 0
            
            # Calculate adherence percentage
            adherence = 0
            if ytd_target > 0:
                # REMOVED: adherence = min((cumulative_coverage / ytd_target) * 100, 100)
                adherence = (cumulative_coverage / ytd_target) * 100  # Allow values above 100%
            
            results.append({
                'month': month,
                'ytd_target': int(ytd_target),
                'ytd_coverage': cumulative_coverage,
                'adherence': round(adherence, 1),
                'monthly_count': monthly_count
            })
        
        return results
        
    except Exception as e:
        print(f"Error calculating month-wise YTD metrics: {str(e)}")
        return []
    finally:
        if conn:
            conn.close()
            
def get_training_wise_metrics(filters):
    """Get metrics for each individual training name including annual target, YTD coverage, and adherence"""
    conn = get_db_connection()
    if not conn:
        return []
    
    try:
        # Get all training names with their categories
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT Training_Name, PMO_Training_Category, PL_Category 
                FROM training_names 
                ORDER BY Training_Name
            """)
            trainings = cursor.fetchall()
        
        results = []
        
        for training in trainings:
            training_name = training['Training_Name']
            pmo_category = training['PMO_Training_Category']
            pl_category = training['PL_Category']
            
            # Create a copy of filters for this specific training
            training_filters = filters.copy()
            training_filters['training_name'] = training_name
            
            # Get metrics for this training
            metrics = calculate_dashboard_metrics(training_filters)
            
            if metrics:
                annual_target = metrics['target_metrics']['target']
                ytd_actual = metrics['ytd_metrics']['ytd_actual']
                ytd_target = metrics['ytd_metrics']['ytd_target']
                adherence = metrics['ytd_metrics']['percentage_adherence']
                
                results.append({
                    'training_name': training_name,
                    'pmo_category': pmo_category,
                    'pl_category': pl_category,
                    'annual_target': annual_target,
                    'ytd_coverage': ytd_actual,
                    'ytd_target': ytd_target,
                    'adherence': adherence
                })
            else:
                # If no metrics available, use zeros
                results.append({
                    'training_name': training_name,
                    'pmo_category': pmo_category,
                    'pl_category': pl_category,
                    'annual_target': 0,
                    'ytd_coverage': 0,
                    'ytd_target': 0,
                    'adherence': 0
                })
        
        return results
        
    except Exception as e:
        print(f"Error getting training-wise metrics: {str(e)}")
        return []
    finally:
        if conn:
            conn.close()
            
def get_employee_group_eor_stats(filters):
    """Get EOR count total and breakdown by employee category and gender"""
    conn = get_db_connection()
    if not conn:
        return {'total_eor_count': 0, 'employee_category_breakdown': {}}
    
    try:
        query = """
            SELECT employee_group, gender, COUNT(DISTINCT per_no) as count
            FROM eor_data WHERE 1=1
        """
        query_params = []
        
        # Apply filters
        if filters.get('factory') and filters['factory'] != 'All':
            query += " AND factory = %s"
            query_params.append(filters['factory'])
        
        if filters.get('bc_no'):
            query += " AND bc_no = %s"
            query_params.append(filters['bc_no'])
            
        if filters.get('gender') and filters['gender'] != 'All':
            query += " AND gender = %s"
            query_params.append(filters['gender'])
        
        query += " GROUP BY employee_group, gender"
        
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            results = cursor.fetchall()
            
            employee_category_breakdown = {}
            total_eor_count = 0
            
            for row in results:
                group = row['employee_group']
                gender = row['gender']
                count = row['count'] or 0
                
                if group not in employee_category_breakdown:
                    employee_category_breakdown[group] = {
                        'total_eor_count': 0,
                        'male_count': 0,
                        'female_count': 0
                    }
                
                employee_category_breakdown[group]['total_eor_count'] += count
                if gender == 'Male':
                    employee_category_breakdown[group]['male_count'] += count
                elif gender == 'Female':
                    employee_category_breakdown[group]['female_count'] += count
                
                total_eor_count += count
            
            return {
                'total_eor_count': total_eor_count,
                'employee_category_breakdown': employee_category_breakdown
            }
            
    except Exception as e:
        print(f"Error getting EOR stats: {str(e)}")
        return {'total_eor_count': 0, 'employee_category_breakdown': {}}
    finally:
        if conn:
            conn.close()

def get_unique_learners_permanent(filters):
    """Get unique learners count for permanent employees only, including EOR count"""
    conn = get_db_connection()
    if not conn:
        return {
            'total_unique_learners': 0, 'male_count': 0, 'female_count': 0,
            'total_eor_count': 0, 'male_eor_count': 0, 'female_eor_count': 0,
            'eor_ytd_target': 0
    }
    try:
        # --- Unique Learners query ---
        query = """
            SELECT gender, COUNT(DISTINCT per_no) as count
            FROM master_data
            WHERE employee_group = 'Permanent'
            AND per_no IS NOT NULL AND per_no != ''
        """
        query_params = []
        query, query_params = apply_standard_filters(query, query_params, filters)
        query += " GROUP BY gender"
        total_unique_learners = male_count = female_count = 0
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            results = cursor.fetchall()
            for row in results:
                gender = row['gender']
                count = row['count'] or 0
                total_unique_learners += count
                if gender == 'Male':
                    male_count += count
                elif gender == 'Female':
                    female_count += count

        # --- EOR query ---
        eor_query = """
            SELECT gender, COUNT(DISTINCT per_no) as count
            FROM eor_data
            WHERE employee_group = 'Permanent'
        """
        eor_params = []
        if filters.get('factory') and filters['factory'] != 'All':
            eor_query += " AND factory = %s"
            eor_params.append(filters['factory'])
        if filters.get('bc_no'):
            eor_query += " AND bc_no = %s"
            eor_params.append(filters['bc_no'])
        if filters.get('gender') and filters['gender'] != 'All':
            eor_query += " AND gender = %s"
            eor_params.append(filters['gender'])
        eor_query += " GROUP BY gender"
        total_eor_count = male_eor_count = female_eor_count = 0
        with conn.cursor() as cursor:
            cursor.execute(eor_query, eor_params)
            eor_results = cursor.fetchall()
            for row in eor_results:
                gender = row['gender']
                count = row['count'] or 0
                total_eor_count += count
                if gender == 'Male':
                    male_eor_count += count
                elif gender == 'Female':
                    female_eor_count += count

        # --- EOR YTD Target calculation ---
        fiscal_month_order = ['April', 'May', 'June', 'July', 'August', 'September',
                              'October', 'November', 'December', 'January', 'February', 'March']

        # Default: use current month if no calendar_month filter provided
        import datetime
        current_month_name = datetime.datetime.now().strftime("%B")

        # Determine which month to use
        month_name = filters.get('calendar_month') or current_month_name
        month_index = 1
        if month_name in fiscal_month_order:
            month_index = fiscal_month_order.index(month_name) + 1

        capped_index = min(month_index, 10)

        if total_eor_count > 0:
            eor_ytd_target = (total_eor_count / 10) * capped_index
        else:
            eor_ytd_target = 0

        return {
            'total_unique_learners': total_unique_learners,
            'male_count': male_count,
            'female_count': female_count,
            'total_eor_count': total_eor_count,
            'male_eor_count': male_eor_count,
            'female_eor_count': female_eor_count,
            'eor_ytd_target': int(eor_ytd_target)
        }
    except Exception as e:
        print(f"Error getting unique learners: {str(e)}")
        return {
            'total_unique_learners': 0, 'male_count': 0, 'female_count': 0,
            'total_eor_count': 0, 'male_eor_count': 0, 'female_eor_count': 0,
            'eor_ytd_target': 0
    }
    finally:
        if conn:
            conn.close()


def get_annual_ytd_metrics(filters):
    """Calculate overall annual target and YTD coverage for the fiscal year"""
    chart_filters = filters.copy()
    chart_filters.pop('calendar_month', None)
    chart_filters.pop('month_range_start', None)
    chart_filters.pop('month_range_end', None)
    chart_filters.pop('month_report_pmo_21_20', None)
    chart_filters.pop('month_cd_key_26_25', None)
    
    dashboard_metrics = calculate_dashboard_metrics(chart_filters)
    annual_target = dashboard_metrics['target_metrics']['target'] if dashboard_metrics else 0
    month_index = get_month_index()
    ytd_coverage = 0
    
    conn = get_db_connection()
    if not conn:
        return {
            'annual_target': annual_target,
            'ytd_coverage': 0,
            'ytd_target': 0,
            'month_index': month_index
        }
    
    try:
        query = """
            SELECT COUNT(id) as count
            FROM master_data
            WHERE 1=1
        """
        query_params = []
        query, query_params = apply_fiscal_year_filter(query, query_params, chart_filters.get('fiscal_year'))
        query, query_params = apply_standard_filters(query, query_params, chart_filters)
        
        with conn.cursor() as cursor:
            cursor.execute(query, query_params)
            result = cursor.fetchone()
            ytd_coverage = result['count'] if result else 0
        
        if annual_target > 0 and month_index > 0:
            ytd_target = (annual_target / 10) * month_index
        else:
            ytd_target = 0
        
        return {
            'annual_target': annual_target,
            'ytd_coverage': ytd_coverage,
            'ytd_target': int(ytd_target),
            'month_index': month_index
        }
        
    except Exception as e:
        print(f"Error calculating annual YTD metrics: {str(e)}")
        return {
            'annual_target': annual_target,
            'ytd_coverage': 0,
            'ytd_target': 0,
            'month_index': month_index
        }
    finally:
        if conn:
            conn.close()